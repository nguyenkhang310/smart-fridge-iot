"""
📊 Export System Parameters to Excel
Quét source code (app.py, firebase_integration.py) và xuất bảng thông số hệ thống ra file Excel.
"""

import re
import os
import sys

# Cài openpyxl nếu chưa có
try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
except ImportError:
    print("⚠ Đang cài openpyxl...")
    os.system(f"{sys.executable} -m pip install openpyxl -q")
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# ── Đường dẫn file nguồn ──
BASE_DIR = os.path.dirname(os.path.abspath(__file__))
APP_PY = os.path.join(BASE_DIR, "app.py")
FIREBASE_PY = os.path.join(BASE_DIR, "core", "firebase_integration.py")
OUTPUT_FILE = os.path.join(BASE_DIR, "ThongSo_HeThong.xlsx")


def read_file(path):
    """Đọc file và trả về list các dòng."""
    with open(path, "r", encoding="utf-8") as f:
        return f.readlines()


def find_value(lines, pattern, cast=str):
    """Tìm giá trị theo regex pattern trong danh sách dòng."""
    for i, line in enumerate(lines):
        m = re.search(pattern, line)
        if m:
            return cast(m.group(1)), i + 1  # (value, line_number)
    return None, None


def extract_params():
    """Trích xuất tất cả thông số hệ thống từ source code."""
    app_lines = read_file(APP_PY)
    fb_lines = read_file(FIREBASE_PY) if os.path.exists(FIREBASE_PY) else []

    params = []

    # Helper: tìm giá trị mặc định trong pattern float(os.environ.get('KEY', 'VALUE'))
    def find_env_default(lines, var_name):
        """Tìm giá trị default của biến dạng: VAR = float(os.environ.get('VAR', 'default'))"""
        for i, line in enumerate(lines):
            # Match cả single và double quotes
            pattern = rf'{var_name}\s*=\s*(?:float|int)\(os\.environ\.get\(["\'][^"\']+["\']\s*,\s*["\']([^"\']+)["\']\)'
            m = re.search(pattern, line)
            if m:
                return m.group(1), i + 1
        return None, None

    # ── 1. ESP32 push lên Firebase (từ firmware, hard-code trong doc) ──
    params.append({
        "thong_so": "ESP32 push lên Firebase",
        "gia_tri": "mỗi 2–3s (/Current)",
        "noi_khai_bao": "Wokwi firmware (file .ino)",
        "giai_thich": "ESP32 gửi dữ liệu cảm biến lên Firebase node /Current mỗi 2–3 giây"
    })

    params.append({
        "thong_so": "ESP32 push History",
        "gia_tri": "mỗi 30s (/History)",
        "noi_khai_bao": "Wokwi firmware (file .ino)",
        "giai_thich": "ESP32 lưu lịch sử cảm biến vào /History mỗi 30 giây"
    })

    # ── 2. Flask poll Firebase ──
    sleep_values = []
    for i, line in enumerate(app_lines):
        if "firebase_update_worker" in line and "def " in line:
            for j in range(i, min(i + 120, len(app_lines))):
                m = re.search(r"time\.sleep\(([\d.]+)\)", app_lines[j])
                if m:
                    sleep_values.append((float(m.group(1)), j + 1))
    
    if sleep_values:
        val = sleep_values[0][0]
        line_no = sleep_values[0][1]
        params.append({
            "thong_so": "Flask poll Firebase",
            "gia_tri": f"mỗi {val}s",
            "noi_khai_bao": f"app.py → firebase_update_worker → time.sleep({val}) (dòng {line_no})",
            "giai_thich": f"Background thread gọi REST API GET đến Firebase mỗi {val} giây"
        })

    # ── 3. SSE gửi về web ──
    for i, line in enumerate(app_lines):
        m = re.search(r"firebase_update_queue\.get\(timeout=([\d.]+)\)", line)
        if m:
            val = float(m.group(1))
            params.append({
                "thong_so": "SSE gửi về web",
                "gia_tri": f"mỗi {val}s",
                "noi_khai_bao": f"app.py → stream_sensors → queue.get(timeout={val}) (dòng {i+1})",
                "giai_thich": f"Dashboard kết nối SSE (/api/sensors/stream), đọc Queue mỗi {val} giây"
            })
            break

    # ── 4. CACHE_MAX_AGE ──
    if fb_lines:
        val, line_no = find_value(fb_lines, r"CACHE_MAX_AGE\s*=\s*(\d+)", int)
        if val is not None:
            params.append({
                "thong_so": "Cache tối đa khi lỗi mạng",
                "gia_tri": f"{val} giây",
                "noi_khai_bao": f"core/firebase_integration.py → CACHE_MAX_AGE = {val} (dòng {line_no})",
                "giai_thich": f"Khi mất kết nối Firebase, dùng dữ liệu cache tối đa {val} giây"
            })

    # ── 5. Camera IP sync interval ──
    val, line_no = find_value(app_lines, r"_CAM_IP_SYNC_INTERVAL\s*=\s*(\d+)", int)
    if val is not None:
        params.append({
            "thong_so": "Sync camera IP từ Firebase",
            "gia_tri": f"mỗi {val}s",
            "noi_khai_bao": f"app.py → _CAM_IP_SYNC_INTERVAL = {val} (dòng {line_no})",
            "giai_thich": f"Polling Firebase mỗi {val} giây để lấy IP mới nhất của ESP32-CAM"
        })

    # ── 6. Door debounce ──
    val, line_no = find_env_default(app_lines, "DOOR_STATE_DEBOUNCE_SECONDS")
    if val is not None:
        params.append({
            "thong_so": "Độ trễ debounce cửa",
            "gia_tri": f"{val}s",
            "noi_khai_bao": f"app.py → DOOR_STATE_DEBOUNCE_SECONDS = {val} (dòng {line_no})",
            "giai_thich": f"Chỉ đổi trạng thái cửa khi tín hiệu ổn định sau {val} giây (chống nhiễu)"
        })

    # ── 7. Door close detect cooldown ──
    val, line_no = find_env_default(app_lines, "DOOR_CLOSE_DETECT_COOLDOWN_SECONDS")
    if val is not None:
        params.append({
            "thong_so": "Cooldown auto-detect khi đóng cửa",
            "gia_tri": f"{val} giây",
            "noi_khai_bao": f"app.py → DOOR_CLOSE_DETECT_COOLDOWN_SECONDS = {val} (dòng {line_no})",
            "giai_thich": f"Sau khi đóng cửa trigger detect, phải chờ {val} giây trước lần tiếp theo"
        })

    # ── 8. Detect dedup window ──
    val, line_no = find_env_default(app_lines, "DETECT_DEDUP_WINDOW_SECONDS")
    if val is not None:
        params.append({
            "thong_so": "Detect dedup window",
            "gia_tri": f"{val} giây",
            "noi_khai_bao": f"app.py → DETECT_DEDUP_WINDOW_SECONDS = {val} (dòng {line_no})",
            "giai_thich": f"Chống trùng lặp detect trong vòng {val} giây"
        })

    # ── 9. Door close trigger delay ──
    val, line_no = find_env_default(app_lines, "DOOR_CLOSE_TRIGGER_DELAY_SECONDS")
    if val is not None:
        params.append({
            "thong_so": "Độ trễ trigger detect sau đóng cửa",
            "gia_tri": f"{val}s",
            "noi_khai_bao": f"app.py → DOOR_CLOSE_TRIGGER_DELAY_SECONDS = {val} (dòng {line_no})",
            "giai_thich": f"Chờ {val} giây sau khi phát hiện đóng cửa mới gọi /api/detect"
        })

    # ── 10. ESP32 stream min interval ──
    val, line_no = find_env_default(app_lines, "ESP32_STREAM_MIN_INTERVAL")
    if val is not None:
        params.append({
            "thong_so": "ESP32 stream khoảng nghỉ tối thiểu",
            "gia_tri": f"{val}s",
            "noi_khai_bao": f"app.py → ESP32_STREAM_MIN_INTERVAL = {val} (dòng {line_no})",
            "giai_thich": f"Khoảng nghỉ tối thiểu {val}s giữa các lần kéo ảnh từ ESP32"
        })

    # ── 11. Door open alert ──
    val, line_no = find_value(app_lines, r"DOOR_OPEN_ALERT_SECONDS\s*=\s*(\d+)", int)
    if val is not None:
        params.append({
            "thong_so": "Cảnh báo cửa mở quá lâu",
            "gia_tri": f"{val} giây",
            "noi_khai_bao": f"app.py → DOOR_OPEN_ALERT_SECONDS = {val} (dòng {line_no})",
            "giai_thich": f"Hiển thị cảnh báo nếu cửa mở liên tục quá {val} giây"
        })

    # ── 12. Brute-force protection ──
    val, line_no = find_env_default(app_lines, "MAX_LOGIN_ATTEMPTS")
    if val is not None:
        params.append({
            "thong_so": "Số lần đăng nhập sai tối đa",
            "gia_tri": f"{val} lần",
            "noi_khai_bao": f"app.py → MAX_LOGIN_ATTEMPTS = {val} (dòng {line_no})",
            "giai_thich": f"Khóa tài khoản sau {val} lần nhập sai liên tiếp"
        })

    val, line_no = find_env_default(app_lines, "LOCKOUT_SECONDS")
    if val is not None:
        v = int(val)
        params.append({
            "thong_so": "Thời gian khóa tài khoản",
            "gia_tri": f"{v} giây ({v//60} phút)",
            "noi_khai_bao": f"app.py → LOCKOUT_SECONDS = {val} (dòng {line_no})",
            "giai_thich": f"Tài khoản bị khóa {v//60} phút sau khi sai quá số lần cho phép"
        })

    # ── 13. Tổng độ trễ ──
    params.append({
        "thong_so": "Tổng độ trễ ESP32 → Web",
        "gia_tri": "~2–4 giây",
        "noi_khai_bao": "Tính từ các bước trên",
        "giai_thich": "ESP32 push (2-3s) + Flask poll (0.2s) + SSE delivery (0.1s) ≈ 2–4 giây"
    })

    return params


def write_excel(params, output_path):
    """Xuất danh sách thông số ra file Excel với định dạng đẹp."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Thông số hệ thống"

    # ── Styles ──
    header_font = Font(name="Arial", size=12, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="2B579A", end_color="2B579A", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

    title_font = Font(name="Arial", size=16, bold=True, color="2B579A")
    subtitle_font = Font(name="Arial", size=10, italic=True, color="666666")

    data_font = Font(name="Arial", size=11)
    data_font_bold = Font(name="Arial", size=11, bold=True)
    value_font = Font(name="Consolas", size=11, bold=True, color="C7254E")
    code_font = Font(name="Consolas", size=10, color="333333")

    even_fill = PatternFill(start_color="F2F6FC", end_color="F2F6FC", fill_type="solid")
    odd_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")

    thin_border = Border(
        left=Side(style="thin", color="CCCCCC"),
        right=Side(style="thin", color="CCCCCC"),
        top=Side(style="thin", color="CCCCCC"),
        bottom=Side(style="thin", color="CCCCCC"),
    )

    data_align = Alignment(vertical="center", wrap_text=True)
    center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # ── Title ──
    ws.merge_cells("A1:D1")
    title_cell = ws["A1"]
    title_cell.value = "📊 Bảng Thông Số Hệ Thống - Smart Fridge IoT"
    title_cell.font = title_font
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 40

    ws.merge_cells("A2:D2")
    sub_cell = ws["A2"]
    sub_cell.value = "Tự động trích xuất từ source code (app.py, firebase_integration.py)"
    sub_cell.font = subtitle_font
    sub_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[2].height = 22

    # ── Headers (row 4) ──
    headers = ["STT", "Thông số", "Giá trị", "Nơi khai báo", "Giải thích"]
    col_widths = [6, 35, 18, 55, 60]
    
    for col_idx, (header, width) in enumerate(zip(headers, col_widths), 1):
        cell = ws.cell(row=4, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border
        ws.column_dimensions[chr(64 + col_idx)].width = width

    ws.row_dimensions[4].height = 30

    # ── Data rows ──
    for row_idx, p in enumerate(params):
        row = row_idx + 5
        fill = even_fill if row_idx % 2 == 0 else odd_fill

        # STT
        cell = ws.cell(row=row, column=1, value=row_idx + 1)
        cell.font = data_font
        cell.fill = fill
        cell.alignment = center_align
        cell.border = thin_border

        # Thông số
        cell = ws.cell(row=row, column=2, value=p["thong_so"])
        cell.font = data_font_bold
        cell.fill = fill
        cell.alignment = data_align
        cell.border = thin_border

        # Giá trị
        cell = ws.cell(row=row, column=3, value=p["gia_tri"])
        cell.font = value_font
        cell.fill = fill
        cell.alignment = center_align
        cell.border = thin_border

        # Nơi khai báo
        cell = ws.cell(row=row, column=4, value=p["noi_khai_bao"])
        cell.font = code_font
        cell.fill = fill
        cell.alignment = data_align
        cell.border = thin_border

        # Giải thích
        cell = ws.cell(row=row, column=5, value=p["giai_thich"])
        cell.font = data_font
        cell.fill = fill
        cell.alignment = data_align
        cell.border = thin_border

        ws.row_dimensions[row].height = 32

    # ── Footer note ──
    footer_row = len(params) + 6
    ws.merge_cells(f"A{footer_row}:E{footer_row}")
    note = ws.cell(row=footer_row, column=1)
    note.value = (
        '💡 Giải thích ngắn gọn: Dữ liệu từ ESP32 được push lên Firebase Realtime Database mỗi 2–3 giây. '
        'Phía server Flask có một background thread (firebase_update_worker) chạy liên tục, cứ 0.2 giây một lần '
        'gọi REST API GET đến Firebase để lấy dữ liệu mới nhất từ node /Current. Khi phát hiện dữ liệu thay đổi, '
        'thread này đẩy vào một Queue nội bộ. Web dashboard kết nối qua SSE (/api/sensors/stream) và đọc Queue '
        'mỗi 0.1 giây để nhận và hiển thị real-time. Tổng độ trễ từ cảm biến ESP32 đến web là khoảng 2–4 giây.'
    )
    note.font = Font(name="Arial", size=10, italic=True, color="555555")
    note.alignment = Alignment(wrap_text=True, vertical="top")
    ws.row_dimensions[footer_row].height = 65

    # ── Freeze header row ──
    ws.freeze_panes = "A5"

    # ── Auto-filter ──
    ws.auto_filter.ref = f"A4:E{len(params) + 4}"

    wb.save(output_path)
    return output_path


if __name__ == "__main__":
    print("🔍 Đang quét source code...")
    params = extract_params()
    print(f"✅ Tìm thấy {len(params)} thông số")

    output = write_excel(params, OUTPUT_FILE)
    print(f"📁 Đã xuất file Excel: {output}")
    print("🎉 Hoàn tất!")
